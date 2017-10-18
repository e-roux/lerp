# -*- coding: utf-8 -*-
"""util.py doctring."""

from collections import namedtuple
from os import mkdir, system
import os.path
import numpy as np
import pandas as pd
from lerp.mesh import mesh2d, mesh3d
from lerp.util.FigureData import go as digitizer

from openpyxl import load_workbook
from openpyxl.utils import rows_from_range
import logging
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger()  # __name__


def to_clipboard_for_excel(array):
    r"""Copy an array into a string format acceptable by Excel.

    Columns separated by \t, rows separated by \n
    """
    try:
        import win32clipboard as clipboard
    except ModuleNotFoundError:
        raise ModuleNotFoundError("Module not found. Not on win32 platform.")
    else:
        # Create string from array
        array_string = ""

        if len(np.shape(array)) == 1:
            array_string += "\t".join([str(_e).replace(".", ",")
                                      for _e in array])

        else:
            for line in array:
                array_string += "\t".join([str(_e).replace(".", ",")
                                          for _e in line])
                array_string += "\r\n"

        # Put string into clipboard (open, clear, set, close)
        clipboard.OpenClipboard()
        clipboard.EmptyClipboard()
        clipboard.SetClipboardText(array_string)
        clipboard.CloseClipboard()


def col2int(n):
    """Conversion utility column string to Integer.

    Returns
    ----------
    Integer representing the column in spreadsheet.

    Note
    ----------
    See get_column_letter() in openpyxl.utils.cell
    """
    from math import pow
    res = 0
    for i, v in enumerate(n.upper()[::-1]):
        res += (ord(v) - 64) * pow(26, i)
    return int(res-1)


class xlsx(object):
    """Generic class for retrieving data from xlsx file.

    Args:
        fileName (str): complete path to xlsx file.

    Attributes:
        fileName (str): complete path to xlsx file.
        wb (str): Human readable string describing the exception.
    """

    def __init__(self, fileName=None):
        """Init."""
        if fileName is not None:
            self.fileName = fileName
            self.wb = load_workbook(fileName, read_only=True, data_only=True)

    def get_range(self, ws=None, r=None, named_range=None, header=True,
                  transpose=False, columns=None, index=None, lerp=False):
        """Fetch data from a xlsx file.

        Parameters
        ----------

            ws : str
                worksheet name where to locate the range r, to also be
                specified in the argument list
            r : str
                range in the workbook in form "A1:Z99"
            named_range : str
                named range in the workbook
            header : bool
                if the column name has to be taken from the
                first line
            transpose : bool
                True if elements are listed row-wise
            columns : str
                columns name
            index : int
                Column index in the original data range

        Returns
        -------
        """
        try:
            # Si named_range est donné en arguement, on récupère
            # le nom de l'onglet et le domaine via la méthode
            # get_named_range
            if named_range is not None:
                _rn = self.wb.get_named_range(named_range).value
                ws, r = [_elt.replace("'", "") for _elt in _rn.split('!')]

            ws = self.wb[ws]
            if transpose is False:
                res = pd.DataFrame([[ws[_c].value for _c in _r]
                                   for _r in rows_from_range(r)])
            else:
                res = pd.DataFrame([[ws[_c].value for _c in _r]
                                   for _r in rows_from_range(r)]).T

            if columns is not None:
                header = False
                res.columns = columns

            if header is True:
                columns = res.iloc[0]
                res.drop(0, inplace=True)
                res.columns = columns

            if type(index) is int:
                _index = res.iloc[:, index]
                res.drop(res.columns[index], axis=1, inplace=True)
                res.index = _index

            if lerp is True:
                if min(res.shape) == 1:
                    if res.shape[0] < res.shape[1]:
                        res = res.T
                    mymesh = mesh2d(res.index, res.iloc[:, 0])
                else:
                    mymesh = mesh3d()
                    mymesh.from_pandas(res)
                return mymesh
            else:
                return res
        except:
            # Tests if the worksheet is present in the xlsx file
            print("{} not in the workboot sheet list".format(ws))

    def to_mesh3d(self):
        """Export to mesh3d."""
        res = mesh3d()
        return res.from_pandas(self)


def quoted_form_of(s: str) -> str:
    """Return a quoted form of a string."""
    return f"\"{s}\""


def save_style(styleName):
    """Save matplotlib configuration into style.

    Info : http://matplotlib.org/users/style_sheets.html
    """
    import codecs
    import matplotlib
    import os
    import matplotlib.pyplot as plt
    from cycler import Cycler
    import numbers

    stylelibDir = '\\'.join([os.path.normpath(matplotlib.get_configdir()),
                            "stylelib"])

    # Check if stylelib path exists
    if not os.access(stylelibDir, os.R_OK):
        os.mkdir(stylelibDir)

    with codecs.open('\\'.join([stylelibDir,
                               os.path.basename(os.path.splitext(styleName)[0])
                                + ".mplstyle "]), "w", "utf-8") as f:

        for k in sorted(plt.rcParams):
            if k.split(".")[0] not in ["animation", "backend", "datapath",
                                       "docstring", "examples", "keymap",
                                       "ps", "svg", "tk", "timezone",
                                       "toolbar", "verbose", "webagg"]:
                v = matplotlib.rcParams[k]

                if v is None:
                    v_rc = "None"
                elif isinstance(v, list):
                    v_rc = ",".join(str(_v) for _v in v)
                elif isinstance(v, tuple):
                    v_rc = "{}".format(v)
                    print(v_rc)
                elif isinstance(v, Cycler):
                    v_rc = "cycler('color', [{}])"\
                                    .format(",".join("'{}'"
                                            .format(str(_v).replace('#', ''))
                                            for _v in v.by_key()['color']))
                else:
                    if isinstance(v, str) and v.startswith("#"):
                        v_rc = v.replace('#', '')
                    else:
                        v_rc = v
                try:
                    plt.rcParams[k] = v_rc
                    f.write(f"{f} : {v_rc}\n")
                except:
                    try:
                        if k == "savefig.bbox":
                            plt.rcParams[k] = 'tight'
                        else:
                            raise
                    except:
                        logger.warning(f"failed on {k} : {v_rc}")


def pdfcrop(fileName):
    """Crop a pdf file inplace.

    Parameters
    ----------
    fileName: str
        Full path to the pdf file to be croped

    Returns
    -------
    np.float64
        The pressure at altitude in bar
    """
    from os.path import normpath

    subprocess.run(["pdfcrop", normpath(fileName), normpath(fileName)])


#  ###### PATH AND FILE RELEATED UTILIES #######
def get_file_list(path=None, extension=None):
    """Get the list of file in the order liked by path."""
    import fnmatch
    import os
    from os.path import basename

    fileList = []
    if isinstance(path, str) or isinstance(path, unicode):
        path = [path]

    if "*" in extension:
        extension = extension
    else:
        extension = '*.{ext}'.format(ext=extension)

    for p in path:
        if p is not None and os.path.exists(p):
            for root, dirnames, filenames in os.walk(p):
                for f in fnmatch.filter(filenames, extension):
                    if not basename(f).startswith('~'):
                        fileList.append(os.path.join(root, f))

    return fileList


def path_name_ext(path):
    """Return a tuple containing root, name ext."""
    out = namedtuple('path', ['root', 'name', 'ext'])

    _p, _e = os.path.splitext(path)
    _r, _n = os.path.split(_p)

    return out(_r, _n, _e)



# Modified from pandas config.py
class DictWrapper(object):
    """ provide attribute-style access to a nested dict"""

    def __init__(self, d, prefix=""):
        object.__setattr__(self, "d", d)
        object.__setattr__(self, "prefix", prefix)

    def __setattr__(self, key, val):
        prefix = object.__getattribute__(self, "prefix")
        if prefix:
            prefix += "."
        prefix += key
        # you can't set new keys
        # you can't overwrite subtrees
        if key in self.d and not isinstance(self.d[key], dict):
            self.d[key] = val
        else:
            raise OptionError("You can only set the value of existing options")

    def __getattr__(self, key):
        prefix = object.__getattribute__(self, "prefix")
        if prefix:
            prefix += "."
        prefix += key
        v = object.__getattribute__(self, "d")[key]
        if isinstance(v, dict):
            return DictWrapper(v, prefix)
        else:
            return v

    def __dir__(self):
        return list(self.d.keys())
