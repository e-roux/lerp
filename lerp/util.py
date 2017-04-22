# -*- coding: utf-8 -*-
"""
    

"""
from collections import namedtuple
from os import mkdir, system
import os.path
import numpy as np
import pandas as pd
from lerp.__init__ import mesh2d, mesh3d

from openpyxl import load_workbook
from openpyxl.utils import rows_from_range
import logging
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger() # __name__

def toClipboardForExcel(array):
    """
    Copies an array into a string format acceptable by Excel.
    Columns separated by \t, rows separated by \n
    """
    import win32clipboard as clipboard

    # Create string from array
    array_string = ""
    
    if len(np.shape(array)) == 1:
        array_string += "\t".join([str(_e).replace(".",",") for _e in array])
        
    else:
        for line in array:
            array_string += "\t".join([str(_e).replace(".",",") for _e in line])
            array_string += "\r\n"

    
    # Put string into clipboard (open, clear, set, close)
    clipboard.OpenClipboard()
    clipboard.EmptyClipboard()
    clipboard.SetClipboardText(array_string)
    clipboard.CloseClipboard()    

class xlsx(object):
    """xlsx

    Generic class for retrieving data from xlsx file

    Args:
        fileName (str): complete path to xlsx file.

    Attributes:
        fileName (str): complete path to xlsx file.
        wb (str): Human readable string describing the exception.
    """        
    def __init__(self, fileName=None):
        if fileName is not None:
            self.fileName = fileName
            self.wb = load_workbook(fileName, read_only=True, data_only=True)

    def get_range(self, ws=None, r=None, named_range=None, header=True,
                  transpose=False, columns = None, index=None, lerp=False):
        """Fetches data from a xlsx file.

        Retrieves ...

        Args:
            ws (str): worksheet name where to locate the range r, to also be
            specified in the argument list
            r (str): range in the workbook in form A1:Z99
            named_range (str) : named range in the workbook
            header (bool): if the column name has to be taken from the
                first line
            transpose (bool): True if elements are listed row-wise 
            columns: columns name
            index (int): Column index in the original data range

        Returns:
            ...

        Raises:
            ...
        """                  
        try:
            # Si named_range est donné en arguement, on récupère
            # le nom de l'onglet et le domaine via la méthode
            # get_named_range
            if named_range is not None:
                _rn = self.wb.get_named_range(named_range).value
                ws, r = [ _elt.replace("'","") for _elt in _rn.split('!')]

            ws = self.wb[ws]
            if transpose is False:
                res = pd.DataFrame([[ws[_c].value for _c in _r] for _r in rows_from_range(r)]) 
            else:
                res = pd.DataFrame([[ws[_c].value for _c in _r] for _r in rows_from_range(r)]).T
            
            if columns is not None:
                header = False
                res.columns = columns
                
            if header is True:
                columns = res.iloc[0]
                res.drop(0, inplace=True)
                res.columns = columns
            
            if type(index) is int:
                _index = res.iloc[:,index]
                res.drop(res.columns[index], axis=1, inplace=True)
                res.index = _index
            
            if lerp is True:
                if min(res.shape) == 1:
                    if res.shape[0] < res.shape[1]:
                        res = res.T
                    myMesh = mesh2d(res.index, res.iloc[:,0])
                else:
                    myMesh = mesh3d()
                    myMesh.from_pandas(res)
                return myMesh
            else:
                return res
        except:
            # Tests if the worksheet is present in the xlsx file
            print("{} not in the workboot sheet list".format(ws))
            
    def to_mesh3d(self):
        res = mesh3d()
        return res.from_pandas(self)
